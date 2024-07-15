import os
from openpyxl import load_workbook
import json

dir_path = 'kb/unit-converter'
unit_files = [f for f in os.listdir(dir_path) if f.endswith(".xlsx")]
units = []
for unit_file in unit_files:
    wb = load_workbook(f'{dir_path}/{unit_file}')
    # Sheets should have the following rows:
    # 1: Measure
    # 2: Decimals
    # 3: Image
    # 4: Default From
    # 5: Default To
    # 6: Units header
    # 7: Label, Factor, Description Header
    # 8: label, factor, description data
    sheet = wb['Sheet1']
    row = 1
    measure = sheet['B'+ str(row)].value
    row += 1
    decimals = sheet['B'+str(row)].value
    row += 1
    image = sheet['B'+str(row)].value
    row += 1
    default_from = sheet['B'+str(row)].value
    row += 1
    default_to = sheet['B'+str(row)].value
    row += 3 # Skip Units header, and Name, Factor, Description header
    unit = dict()
    unit['measure'] = measure
    unit['decimals'] = decimals
    unit['image'] = image
    unit['units'] = []
    print(measure)
    print(decimals)
    print(image)
    label = sheet.cell(row=row, column=1).value
    factor = sheet.cell(row=row, column=2).value
    description = sheet.cell(row=row, column=3).value
    while (label is not None):
        u = dict()
        u['label'] = label
        u['factor'] = factor
        u['description'] = description
        if row == default_from:
            unit['default_from'] = u
        if row == default_to:
            unit['default_to'] = u
        unit['units'].append(u)
        print(label, factor)
        row += 1
        label = sheet.cell(row=row, column=1).value
        factor = sheet.cell(row=row, column=2).value
        description = sheet.cell(row=row, column=3).value
    units.append(unit)

with open('generate/static/unit-conversions.json', 'w') as fp:
    output=json.dumps(units, indent=4)
    fp.write(output)
